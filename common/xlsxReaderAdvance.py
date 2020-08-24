# -*- coding: utf-8 -*-
import sys, os
import copy
from convert import convert
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import *
from openpyxl.styles import *
from openpyxl.styles.colors import Color
#from openpyxl.drawing import *
reload(sys)
sys.setdefaultencoding('utf8')

from datetime import *

class xlsxReaderAdvance(object):
    def __init__(self):
        self.conObj = convert() 

    def get_mergecells_dict(self, merged_cells):
        mdict = {}
        for mcell in merged_cells:
            rcells = range_boundaries(mcell)
            mdict[(rcells[1], rcells[0])] = (rcells[3], rcells[2])
            
        return mdict

    def get_cdict(self):
        cdict = {'data' :'',
                'formula': '',
                'formatStr': '',
                'comment_dict':{},
                'colspan':1,
                'rowspan':1,
                'font_dict': {},
                'border_dict':{},
                'fill_dict':{},
                'align_dict':{}
                } 
        return cdict

    def getColorDict(self, colrObj):
        if not colrObj:
            return {}
        
        colordict = {}
        colordict['tint'] = colrObj.tint
        if type( colrObj.auto) == bool:
            colordict['auto'] = colrObj.auto
        else:
            colordict['auto'] = False
        if type(colrObj.theme) == long:
            colordict['theme'] = colrObj.theme
        else:
            colordict['theme'] = None
        if type(colrObj.rgb) == str:
            colordict['rgb'] = colrObj.rgb
        else:
            colordict['rgb'] = None
        if type(colrObj.indexed) == long:
            colordict['indexed'] = colrObj.indexed
        else:
            colordict['indexed'] = None
        colordict['type'] = colrObj.type
        return colordict

    def getFontDict(self, fontObj):
        if not fontObj: return {}
        fontdict = {}
        fontdict['name'] = fontObj.name
        fontdict['charset'] = fontObj.charset
        fontdict['family'] = fontObj.family
        fontdict['b'] = fontObj.b
        fontdict['i'] = fontObj.i
        fontdict['u'] = fontObj.u
        fontdict['strike'] = fontObj.strike
        fontdict['outline'] = fontObj.outline
        fontdict['shadow'] = fontObj.shadow
        fontdict['condense'] = fontObj.condense
        fontdict['extend'] = fontObj.extend
        fontdict['sz'] = fontObj.sz
        fontdict['vertAlign'] = fontObj.vertAlign
        fontdict['scheme'] = fontObj.scheme
        fontdict['colordict'] = self.getColorDict(fontObj.color)
        return fontdict

    def getBorderStyleDict(self, styleObj):
        if not styleObj: return {}
        styledict = {}
        colordict = self.getColorDict(styleObj.color)
        styledict['colordict'] = colordict
        styledict['style'] = styleObj.style        
        return styledict


    def getBorderDict(self, borderObj):
        if not borderObj: return {}
        border_dict = {}
        border_dict['outline'] = borderObj.outline
        border_dict['diagonalUp'] = borderObj.diagonalUp
        border_dict['diagonalDown'] = borderObj.diagonalDown
        border_dict['left'] = self.getBorderStyleDict(borderObj.left)
        border_dict['right'] = self.getBorderStyleDict(borderObj.right)
        border_dict['top'] = self.getBorderStyleDict(borderObj.top)
        border_dict['bottom'] = self.getBorderStyleDict(borderObj.bottom)
        border_dict['diagonal'] = self.getBorderStyleDict(borderObj.diagonal)
        border_dict['vertical'] = borderObj.vertical
        border_dict['horizontal'] = borderObj.horizontal
        return border_dict

    def getFillDict(self, fillObj):
        if not fillObj: return {}
        fill_dict = {}
        fill_dict['patternType'] = fillObj.patternType
        fill_dict['fgColor'] = self.getColorDict(fillObj.fgColor)
        fill_dict['bgColor'] = self.getColorDict(fillObj.bgColor)
        return fill_dict

    def getAlignmentDict(self, alignObj):
        if not alignObj: return {}
        align_dict = {}
        align_dict['indent'] = alignObj.indent
        align_dict['vertical'] = alignObj.vertical
        align_dict['relativeIndent'] = alignObj.relativeIndent
        align_dict['shrinkToFit'] = alignObj.shrinkToFit
        align_dict['justifyLastLine'] = alignObj.justifyLastLine
        align_dict['readingOrder'] = alignObj.readingOrder
        align_dict['wrapText'] = alignObj.wrapText
        align_dict['horizontal'] = alignObj.horizontal
        align_dict['textRotation'] = alignObj.textRotation
        return align_dict

    def getCommentDict(self, commentObj):
        comment_dict = {}
        comment_dict['author'] = commentObj.author
        #comment_dict['bind'] = commentObj.bind
        comment_dict['content'] = commentObj.content
        comment_dict['height'] = commentObj.height
        #comment_dict['parent'] = commentObj.parent
        comment_dict['text'] = commentObj.text
        #comment_dict['unbind'] = commentObj.unbind
        comment_dict['width'] = commentObj.width
        return comment_dict

    def getGeographicalPropDict(self, geoPropObj):
        geoprop_dict = {}
         

        return geoprop_dict

    def getChartDict(self, chartObj):
        chart_dict = {}

        return chart_dict

    def readExcel(self, fname):
        excel_data = {} 
        data_wb = load_workbook(filename=fname, data_only=True)
        wb = load_workbook(filename=fname, data_only=False)
        visible_sheets = [idx for idx, sheet in enumerate(wb._sheets) if sheet.sheet_state == "visible"]
        sheet_names = wb.get_sheet_names()
        for idx, sheet_name in enumerate(sheet_names):
            sheetObj = wb.get_sheet_by_name(sheet_name)

            #print '---------------------------------------------------'
            #print sheet_name
            #print dir(sheetObj)

            #ROW HEIGHT
            row_outline_dict, rowHeights_dict = {}, {}
            for i in range(sheetObj.max_row):
                ob = sheetObj.row_dimensions[i+1]
                if ob.height: rowHeights_dict[i+1] = ob.height
                if ob.outline_level: 
                    if not row_outline_dict.get(i+1):
                        row_outline_dict[i+1] = []
                    row_outline_dict[i+1].append({'outline_level': ob.outline_level, 'hidden': ob.hidden})

            #COLUMN WIDTH
            col_outline_dict, colWidths_dict = {}, {}
            for i in range(sheetObj.max_column):
                ob = sheetObj.column_dimensions[get_column_letter(i+1)]
                if ob.width: colWidths_dict[get_column_letter(i+1)] = ob.width
                if ob.outline_level: 
                    if not col_outline_dict.get(i+1):
                        col_outline_dict[i+1] = []
                    col_outline_dict[i+1].append({'outline_level':ob.outline_level, 'hidden': ob.hidden})
                    
            #GRIDLINES
            gridlines = sheetObj.sheet_view.showGridLines
          
            #FREEZE_PANES 
            freeze_panes = sheetObj.freeze_panes
            
            merged_cell_ranges = sheetObj.merged_cell_ranges
            mdict = self.get_mergecells_dict(merged_cell_ranges)
            sheet_name = str(sheet_name)
            excel_data[sheet_name] = {}
            excel_data[sheet_name]['visible'] = 1 if idx in visible_sheets else 0
            excel_data[sheet_name]['sheet_order'] = idx
            excel_data[sheet_name]['column_widths_dict'] = colWidths_dict
            excel_data[sheet_name]['row_heights_dict'] = rowHeights_dict
            excel_data[sheet_name]['gridlines'] = gridlines
            excel_data[sheet_name]['freeze_panes'] = freeze_panes
            excel_data[sheet_name]['merged_cell_ranges'] = merged_cell_ranges
            excel_data[sheet_name]['col_outline_dict'] = col_outline_dict
            excel_data[sheet_name]['row_outline_dict'] = row_outline_dict
            excel_data[sheet_name]['sheet_name'] = sheet_name
            celldict = {}
            mrow, mcol = 0, 0
            for rowid, rowObjs in enumerate(sheetObj.rows, 1):
                mrow = max(mrow, rowid)
                #dataSheetObj_rows = list(dataSheetObj.rows) 
                for colidx, cellObj in enumerate(rowObjs):
                    colid = colidx + 1
                    mcol = max(mcol, colid)
                    #print cellObj
                    #print dir(cellObj)  
                    formula = ''
                    try:
                        formula = cellObj.value
                    except:
                        #print 'ERROR IN VAL - ', sheet_name, rowid, colid, cellObj 
                        pass    
                    #print sheet_name, rowid, colid, cellObj.value
                    formula = cellObj.value
                    #print sheet_name, "=====", rowid, colid, formula
                    if str(formula) and str(formula)[0] != '=': formula = ''
                    formatStr = cellObj.number_format
                    formatStr = formatStr.replace('\\', '') 
                    font_dict, border_dict, fill_dict, align_dict = {}, {}, {}, {}
                    comment_dict = {}
                    if cellObj.has_style:
                        font_dict = self.getFontDict(cellObj.font)
                        border_dict = self.getBorderDict(cellObj.border)
                        fill_dict = self.getFillDict(cellObj.fill)
                        align_dict = self.getAlignmentDict(cellObj.alignment)
                        #print '>>>> ', font_dict, border_dict, fill_dict, align_dict
                    if cellObj.comment:
                        comment_dict = self.getCommentDict(cellObj.comment)
                        
                    data = ''
                    cell = (rowid, colid)
                    cdict = self.get_cdict()
                    cdict['formula'] = formula
                    cdict['comment_dict'] = comment_dict
                    cdict['formatStr'] = formatStr
                    cdict['font_dict'] = font_dict
                    cdict['border_dict'] = border_dict
                    cdict['fill_dict'] = fill_dict
                    cdict['align_dict'] = align_dict
                    if mdict.get(cell, ()):
                        rspan = mdict[cell][0] - rowid
                        cspan = mdict[cell][1] - colid
                        cdict['rowspan'] = rspan + 1
                        cdict['colspan'] = cspan + 1

                    celldict[cell] = cdict
                excel_data[sheet_name]['celldict'] = celldict
                excel_data[sheet_name]['rows'] = mrow
                excel_data[sheet_name]['cols'] = mcol

        #Adding data, it include calculated data
        for idx, sheet_name in enumerate(sheet_names):
            dataSheetObj = data_wb.get_sheet_by_name(sheet_name)
            sheet_name = str(sheet_name)
            for rowid, rowObjs in enumerate(dataSheetObj.rows, 1):
                for colidx, cellObj in enumerate(rowObjs):
                    colid = colidx + 1
                    data = cellObj.value
                    if data == None : data = ''
                    try: data = str(data)
                    except: pass
                    data = ' '.join(map(lambda data:data.strip(), data.split()))

                    cell = (rowid, colid)
                    if not excel_data[sheet_name]['celldict'].get(cell, {}):
                        excel_data[sheet_name]['celldict'][cell] = self.get_cdict()
                    
                    excel_data[sheet_name]['celldict'][cell]['data'] = data

        return  excel_data

    def process(self, dataPath, file_name, data_flg=False, comment_flg=False):
        iFile = os.path.join(dataPath, file_name)
        if not os.path.exists(iFile): return {}
        print 'Reading file : ', iFile
        excel_op_data = self.readExcel(iFile)
        return excel_op_data

#TEST
if __name__=="__main__":
    obj = xlsxReader()
    dataPath = sys.argv[1]
    caseID = sys.argv[2] 
    obj.process(dataPath, caseID)
